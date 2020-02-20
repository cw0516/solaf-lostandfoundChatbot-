# excel에서 사용자의 세션 정보 관련 모듈

from openpyxl import load_workbook

EXCEL_FILE_NAME = "USER.xlsx"
db = load_workbook(filename=EXCEL_FILE_NAME)
user_db = db['user_info']

def get_user(user_id, user_name):
    '''
        USER.xlsx에서
        기존 챗본 사용자라면 행번호를 반환
        없으면 DB에 업데이트하고 행번호 반환
    '''
    user_exist = False

    for row in user_db.rows:

        if row[0].value == user_id:

            user_exist = True
            user_row = row[0].row
            db.save(EXCEL_FILE_NAME)
            break  # return user_row

    if not user_exist:

        user_db[user_db.max_row + 1][0].value = user_id
        user_db[user_db.max_row][1].value = user_name
        user_db[user_db.max_row][2].value = None #CATEGORY
        user_db[user_db.max_row][3].value = None #DESCRIPTION
        user_db[user_db.max_row][4].value = None #DATE
        user_db[user_db.max_row][5].value = None #PLACE
        user_db[user_db.max_row][6].value = 'DATE' #STAGE
        user_db[user_db.max_row][7].value = None  # DATE_QUERY
        user_db[user_db.max_row][8].value = None  # DATE_QUERY
        user_db[user_db.max_row][9].value = None  # DATE_QUERY
        user_row = user_db.max_row
        db.save(EXCEL_FILE_NAME)

    return user_row



def reset_with_start(user_row):

    #userid와 name은 지울 필요없다
    for i in range(2,10):
        if i == 6:
            user_db[user_row][i].value = 'DATE'
        else:
            user_db[user_row][i].value = None

    db.save(EXCEL_FILE_NAME)

#date
def save_date(user_row,date):
    user_db[user_row][7].value = date
    db.save(EXCEL_FILE_NAME)
    return

def get_date(user_row):
    return user_db[user_row][7].value

#step
def save_step(user_row, status):
    if status == 'DESCRIPTION':
        user_db[user_row][6].value = 'DESCRIPTION'
        db.save(EXCEL_FILE_NAME)

def get_step(user_row):
    return user_db[user_row][6].value

#description
def save_description(user_row,description):
    user_db[user_row][3].value = description
    db.save(EXCEL_FILE_NAME)
    return

def get_description(user_row):
    return user_db[user_row][3].value


# center
def save_center(user_row,center):

    user_db[user_row][9].value = center
    db.save(EXCEL_FILE_NAME)
    return

def get_center(user_row):
    return user_db[user_row][9].value


# center_num
def save_center_num(user_row,center_num):
    user_db[user_row][8].value = center_num
    db.save(EXCEL_FILE_NAME)
    return

def get_center_num(user_row):

    return user_db[user_row][8].value


# center_code
def save_code(user_row,code):
    user_db[user_row][10].value = code
    db.save(EXCEL_FILE_NAME)
    return

def get_code(user_row):
    return user_db[user_row][10].value

